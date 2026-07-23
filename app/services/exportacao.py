from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from app.database.models import Agendamento


COLUNAS = [
    "Data",
    "Horário",
    "Paciente",
    "CPF",
    "Médico",
    "Especialidade",
    "Convênio",
    "Status",
]


def _linhas_agendamentos(
    agendamentos: list[Agendamento],
) -> list[list[str]]:
    linhas = []

    for item in agendamentos:
        paciente = item.paciente

        linhas.append(
            [
                item.data.strftime("%d/%m/%Y"),
                item.horario.strftime("%H:%M"),
                paciente.nome if paciente else "",
                paciente.cpf if paciente else "",
                item.medico,
                item.especialidade,
                item.convenio or "Particular",
                item.status,
            ]
        )

    return linhas


def gerar_excel(
    agendamentos: list[Agendamento],
) -> BytesIO:
    buffer = BytesIO()
    workbook = Workbook()
    planilha = workbook.active
    planilha.title = "Agendamentos"

    cabecalho_fill = PatternFill(
        start_color="0D6EFD",
        end_color="0D6EFD",
        fill_type="solid",
    )
    cabecalho_font = Font(
        bold=True,
        color="FFFFFF",
    )

    planilha.append(COLUNAS)

    for coluna in range(1, len(COLUNAS) + 1):
        celula = planilha.cell(
            row=1,
            column=coluna,
        )
        celula.fill = cabecalho_fill
        celula.font = cabecalho_font
        celula.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

    for linha in _linhas_agendamentos(agendamentos):
        planilha.append(linha)

    for coluna in planilha.columns:
        largura_max = 0
        coluna_letra = coluna[0].column_letter

        for celula in coluna:
            valor = str(celula.value or "")
            largura_max = max(largura_max, len(valor))

        planilha.column_dimensions[coluna_letra].width = min(
            largura_max + 2,
            40,
        )

    workbook.save(buffer)
    buffer.seek(0)

    return buffer


def gerar_pdf(
    agendamentos: list[Agendamento],
) -> BytesIO:
    buffer = BytesIO()
    documento = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=1.5 * cm,
        rightMargin=1.5 * cm,
        topMargin=1.5 * cm,
        bottomMargin=1.5 * cm,
    )

    estilos = getSampleStyleSheet()
    titulo = ParagraphStyle(
        "TituloRelatorio",
        parent=estilos["Heading1"],
        fontSize=16,
        spaceAfter=12,
    )
    subtitulo = ParagraphStyle(
        "SubtituloRelatorio",
        parent=estilos["Normal"],
        fontSize=10,
        textColor=colors.grey,
        spaceAfter=16,
    )

    elementos = [
        Paragraph(
            "Relatório de Agendamentos",
            titulo,
        ),
        Paragraph(
            f"Total de registros: {len(agendamentos)}",
            subtitulo,
        ),
        Spacer(
            1,
            0.3 * cm,
        ),
    ]

    dados = [COLUNAS, *_linhas_agendamentos(agendamentos)]

    if len(dados) == 1:
        dados.append(["Nenhum agendamento encontrado"] + [""] * (len(COLUNAS) - 1))

    tabela = Table(
        dados,
        repeatRows=1,
    )

    tabela.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0D6EFD")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.lightgrey),
                (
                    "ROWBACKGROUNDS",
                    (0, 1),
                    (-1, -1),
                    [colors.white, colors.HexColor("#F8F9FA")],
                ),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )

    elementos.append(tabela)
    documento.build(elementos)
    buffer.seek(0)

    return buffer
